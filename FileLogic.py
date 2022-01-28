from GUI import *
from tkinter import messagebox
from openpyxl import *
from tkinter import *


class logic:

    def GetExcelFile(self, final: str, reportname,excelName = "ExcelNew"):
        #print(os.path.exists("ExcelNew.xlsx"))
        ExcelFileName = excelName+".xlsx"
        if os.path.exists(ExcelFileName):
            wb = load_workbook(ExcelFileName)
            ws = wb.active
            data = [(reportname, final, 'NA', 'NA')]
            for i in data:
                ws.append(i)
        else:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Report Name"
            ws["B1"] = "Path of Generated reports"
            ws["C1"] = "Path of Standard reports"
            ws["D1"] = "Result"
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 10
            data = [(reportname, final, 'NA', 'NA')]
            for j in data:
                ws.append(j)
        wb.save(ExcelFileName)

    def FileExecution(self, source, destination, fName, cName):

        if len(source) != 0 and len(destination) != 0:
            try:
                SOURCE = source
                name = SOURCE.split("/")
                MIDDLE = "/Users/srichanakyachowdary/Desktop/GeneratedFiles1/newFolder/" + name[len(name) - 1]
                os.renames(SOURCE, MIDDLE)
                #time.sleep(5)
                DESTINATION = destination
                FILENAME = DESTINATION + "/" + fName + "-" + cName + ".pdf"
                initial = MIDDLE
                final = FILENAME
                os.rename(initial, final)
                #print(excelName)
                self.GetExcelFile(FILENAME, fName)
                messagebox.showinfo(title="Success", message="File copied from source to destination")
            except:
                messagebox.showerror(title="Error", message="File Missing")

        else:
            messagebox.showerror(title="Error!!!", message="File path missing")

